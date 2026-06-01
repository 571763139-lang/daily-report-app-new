import os
import re
import io
import uuid
import json
import warnings
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file, render_template

import pandas as pd
import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

app = Flask(__name__, template_folder='.', static_folder='static')

@app.before_request
def handle_options():
    if request.method == 'OPTIONS':
        return app.make_default_options_response()

@app.after_request
def after_request(response):
    response.headers.pop('Access-Control-Allow-Origin', None)
    response.headers.pop('Access-Control-Allow-Headers', None)
    response.headers.pop('Access-Control-Allow-Methods', None)
    
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response


# 内存文件缓存字典，用于暂存生成的 Excel 文件流
# 格式为: { file_id: { "filename": str, "data": BytesIO } }
GENERATED_FILES = {}

# 固定模板名称
TEMPLATE_FILE = "新模板.xlsx"


def process_weighing_data(weighing_file_stream):
    """计算污水外运量数据，接收文件流"""
    try:
        df = pd.read_excel(weighing_file_stream)
        # 过滤掉行政区域为白云区的数据，且收集者包含环投环境
        df_filtered = df[df['行政区域'].fillna('') != '白云区']
        df_filtered = df_filtered[df_filtered['收集者'].fillna('').str.contains('环投环境')]

        total = df_filtered['垃圾重量'].sum()
        to_hf = df_filtered[df_filtered['货物去向'].fillna('').str.contains('兴丰')]['垃圾重量'].sum()
        to_power = total - to_hf
        return round(total, 2), round(to_hf, 2), round(to_power, 2)
    except Exception as e:
        print(f"❌ 读取称重数据失败: {e}")
        return 0.0, 0.0, 0.0


def extract_float(pattern, text):
    """安全提取文字中的数字"""
    match = re.search(pattern, text)
    return float(match.group(1)) if match else 0.0


def get_row_by_keywords(sheet, keywords_list, default_row):
    """获取区域分界行号"""
    for r in range(1, sheet.max_row + 1):
        row_text = "".join(
            [str(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1) if sheet.cell(r, c).value])
        row_text = row_text.replace(' ', '').replace('\n', '')
        for kw in keywords_list:
            if kw in row_text:
                return r
    return default_row


def inject_manual_data(sheet, keywords_list, value, start_row=1, end_row=None):
    """【网格数据注入】寻找周围独立的 0 并替换"""
    max_r = end_row or sheet.max_row
    for r in range(start_row, max_r + 1):
        row_text = "".join(
            [str(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1) if sheet.cell(r, c).value])
        row_text = row_text.replace(' ', '').replace('\n', '')

        for kw in keywords_list:
            if kw in row_text:
                for search_r in range(r, min(r + 4, sheet.max_row + 1)):
                    for c in range(1, sheet.max_column + 1):
                        cell = sheet.cell(search_r, c)
                        if str(cell.value).strip() in ['0', '0.0', '0吨', '0方', '0.00']:
                            cell.value = value
                            return True
    return False


def copy_system_data_sequential(raw_sheet, tmpl_sheet):
    """【系统数据精准注入】完美解决数字存为文本及标号穿透问题"""
    # 1. 拷贝日期标题
    title_val = None
    for r in range(1, 4):
        for c in range(1, 5):
            val = str(raw_sheet.cell(r, c).value)
            if "生产运营情况" in val:
                title_val = val
                break
        if title_val:
            break

    if title_val:
        for r in range(1, 4):
            for c in range(1, 5):
                if "生产运营情况" in str(tmpl_sheet.cell(r, c).value):
                    tmpl_sheet.cell(r, c).value = title_val
                    break

    # 2. 从上往下顺藤摸瓜寻找对应项
    tmpl_search_start = 1
    for raw_r in range(1, raw_sheet.max_row + 1):
        numbers, texts = [], []

        for c in range(1, raw_sheet.max_column + 1):
            val = raw_sheet.cell(raw_r, c).value
            if val is None:
                continue

            if isinstance(val, (int, float)):
                numbers.append(val)
            elif isinstance(val, str):
                clean_str = val.strip()
                try:
                    num_val = float(clean_str.replace(',', ''))
                    numbers.append(num_val)
                except ValueError:
                    clean_str_no_space = clean_str.replace(' ', '').replace('\n', '')
                    if len(clean_str_no_space) > 1:
                        texts.append(clean_str_no_space)

        if numbers and texts:
            best_label = texts[-1]

            # 提取文本开头的序号
            p_raw = re.match(r'^(\d+(?:\.\d+)*)[-－—_~、.]', best_label)
            raw_prefix = p_raw.group(1) if p_raw else None

            found_tmpl_r = -1
            for t_r in range(tmpl_search_start, tmpl_sheet.max_row + 1):
                row_match = False
                for c in range(1, tmpl_sheet.max_column + 1):
                    t_val_raw = tmpl_sheet.cell(t_r, c).value
                    if not isinstance(t_val_raw, str):
                        continue

                    t_val = t_val_raw.replace(' ', '').replace('\n', '')
                    if len(t_val) < 2:
                        continue

                    if raw_prefix:
                        p_tmpl = re.match(r'^(\d+(?:\.\d+)*)[-－—_~、.]', t_val)
                        if p_tmpl and p_tmpl.group(1) == raw_prefix:
                            row_match = True
                            break
                    else:
                        if best_label in t_val or t_val in best_label:
                            row_match = True
                            break

                if row_match:
                    found_tmpl_r = t_r
                    break

            if found_tmpl_r != -1:
                num_idx = 0
                for c in range(1, tmpl_sheet.max_column + 1):
                    cell = tmpl_sheet.cell(found_tmpl_r, c)
                    if str(cell.value).strip() in ['0', '0.0', '0.00']:
                        cell.value = numbers[num_idx]
                        num_idx += 1
                        if num_idx >= len(numbers):
                            break
                tmpl_search_start = found_tmpl_r + 1


def fix_bottom_paragraphs(raw_sheet, tmpl_sheet, tot_sewage, to_hf, to_pwr, hf_data, checks):
    """【表尾段落文本内手术】包含检查项的填报与污水转运数据的替换"""
    boiler_str, turbine_str, date_str = None, None, ""

    # 1. 解析日期标题
    for c in range(1, 5):
        val = str(raw_sheet.cell(1, c).value)
        if "生产运营情况" in val:
            match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', val)
            if match:
                date_str = f"{int(match.group(2))}月{int(match.group(3))}日"
            break

    # 2. 获取锅炉/汽机正常数
    for r in range(1, raw_sheet.max_row + 1):
        for c in range(1, raw_sheet.max_column + 1):
            val = str(raw_sheet.cell(r, c).value)
            if "锅炉总数" in val:
                boiler_str = raw_sheet.cell(r, c).value
            if "汽机总数" in val:
                turbine_str = raw_sheet.cell(r, c).value

    # 3. 构造检查部分的替换字符串
    if checks:
        num_symbols = ["①", "②", "③", "④", "⑤", "⑥"]
        check_lines = []
        for i, chk in enumerate(checks[:6]):
            val_strip = str(chk).strip()
            if val_strip and val_strip != "无":
                check_lines.append(f"{num_symbols[i]}.{val_strip}")
        if check_lines:
            check_text = "1.检查：\n" + "\n".join(check_lines)
        else:
            check_text = "1.检查：\n  无。"
    else:
        check_text = "1.检查：\n  无。"

    # 4. 表尾单元格遍历更新
    for r in range(1, tmpl_sheet.max_row + 1):
        for c in range(1, tmpl_sheet.max_column + 1):
            cell = tmpl_sheet.cell(r, c)
            val = cell.value

            if isinstance(val, str):
                if "锅炉总数" in val and boiler_str:
                    cell.value = boiler_str
                    continue
                if "汽机总数" in val and turbine_str:
                    cell.value = turbine_str
                    continue

                new_val = val
                changed = False

                # 替换检查内容
                if "1.检查：" in new_val:
                    # 使用正则匹配 "1.检查：" 到 "2.垃圾转运站" 之间的任意文本，并替换
                    new_val = re.sub(r'1\.检查：[\s\S]*?(?=2\.垃圾转运站)', check_text + "\n", new_val)
                    changed = True

                # 替换污水转运文本
                if "污水" in new_val or "转运" in new_val or "环服" in new_val:
                    if date_str:
                        new_val = re.sub(r'[0O零]+月[0O零]+日', date_str, new_val)

                    rules = [
                        (r'(污水外运量共计[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{tot_sewage}\\2'),
                        (r'(其中去环服[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{to_hf}\\2'),
                        (r'(去电厂[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{to_pwr}\\2'),
                        (r'(污水共计[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{hf_data.get("总量", 0)}\\2'),
                        (r'(接收环境集团.*?[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{to_hf}\\2'),
                        (r'(白云区[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{hf_data.get("白云区", 0)}\\2'),
                        (r'(人和镇[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{hf_data.get("人和镇", 0)}\\2'),
                        (r'(天河城管局[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{hf_data.get("天河城管", 0)}\\2'),
                        (r'(九龙镇[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{hf_data.get("九龙镇", 0)}\\2'),
                        (r'(太和镇[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{hf_data.get("太和镇", 0)}\\2'),
                        (r'(云埔街道[：:]?\s*)[0\.]+(\s*吨)', f'\\g<1>{hf_data.get("云埔街道", 0)}\\2'),
                    ]
                    for pattern, repl in rules:
                        new_val = re.sub(pattern, repl, new_val)
                    changed = True

                if changed:
                    cell.value = new_val


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/generate', methods=['POST'])
def generate_report():
    try:
        # 1. 获取文件及表单数据
        mis_file = request.files.get('mis_file')
        weighing_file = request.files.get('weighing_file')
        baiyun_text = request.form.get('baiyun_text', '')
        zengcheng_text = request.form.get('zengcheng_text', '')
        huanfu_text = request.form.get('huanfu_text', '')
        checks_json = request.form.get('checks', '[]')

        if not mis_file or not weighing_file:
            return jsonify({'success': False, 'error': '请上传 MIS数据 和 称重数据 文件！'}), 400

        try:
            checks = json.loads(checks_json)
        except Exception:
            checks = []

        # 2. 检查内置模板是否存在
        if not os.path.exists(TEMPLATE_FILE):
            return jsonify({'success': False, 'error': f'系统内置的模板文件 `{TEMPLATE_FILE}` 丢失，请检查服务端文件配置。'}), 500

        # 3. 处理称重数据获取污水外运数据
        tot_sewage, to_hf, to_pwr = process_weighing_data(weighing_file)

        # 4. 解析白云建废文字数据
        by_shizha = extract_float(r'石渣）?[^\d]*([\d.]+)', by_run_text := baiyun_text)
        by_shuinitou = extract_float(r'水泥头）?[^\d]*([\d.]+)', baiyun_text)
        by_yuanshengshi = extract_float(r'原生石）?[^\d]*([\d.]+)', baiyun_text)
        by_jianfei_total = by_shizha + by_shuinitou + by_yuanshengshi

        by_hunningtu = extract_float(r'原生混凝土[^\d]*([\d.]+)', baiyun_text)
        by_zaisheng12 = extract_float(r'再生1-2石[^\d]*([\d.]+)', baiyun_text)
        by_zaishengshuiwen = extract_float(r'再生水稳[^\d]*([\d.]+)', baiyun_text)

        # 5. 解析增城建废文字数据 (新改动：自动识别文本)
        z_zx = extract_float(r'装修垃圾(?:进厂量)?[^\d]*([\d.]+)', zengcheng_text)
        z_sw = extract_float(r'再生水稳(?:量)?[^\d]*([\d.]+)', zengcheng_text)
        z_sf = extract_float(r'再生石粉(?:量)?[^\d]*([\d.]+)', zengcheng_text)
        z_zz = extract_float(r'砖渣[^\d]*([\d.]+)', zengcheng_text)
        z_hnt = extract_float(r'混凝土(?:块)?[^\d]*([\d.]+)', zengcheng_text)

        # 6. 解析环服接收污水数据
        hf_data = {
            '总量': extract_float(r'进场总量[^\d]*([\d.]+)', huanfu_text),
            '白云区': extract_float(r'白云区[^\d]*([\d.]+)', huanfu_text),
            '人和镇': extract_float(r'人和镇[^\d]*([\d.]+)', huanfu_text),
            '太和镇': extract_float(r'太和镇[^\d]*([\d.]+)', huanfu_text),
            '云埔街道': extract_float(r'云埔街道[^\d]*([\d.]+)', huanfu_text),
            '九龙镇': extract_float(r'九龙镇[^\d]*([\d.]+)', huanfu_text),
            '天河城管': extract_float(r'天河城管局[^\d]*([\d.]+)', huanfu_text),
        }

        # 7. 数据校验比对
        calc_hf_sum = to_hf + hf_data['白云区'] + hf_data['人和镇'] + hf_data['太和镇'] + \
                      hf_data['云埔街道'] + hf_data['九龙镇'] + hf_data['天河城管']

        check_status = "success"
        check_message = "环服公司接收污水数量相加核对无误！"
        if abs(calc_hf_sum - hf_data['总量']) >= 0.01:
            check_status = "warning"
            check_message = f"环服公司接收污水数量有误！(各项相加={round(calc_hf_sum, 2)}, 但文本总量={hf_data['总量']})"

        # 8. 读取工作簿进行填充
        tmpl_wb = openpyxl.load_workbook(TEMPLATE_FILE)
        tmpl_sheet = tmpl_wb.active
        raw_wb = openpyxl.load_workbook(mis_file)
        raw_sheet = raw_wb.active

        # 动态提取报告日期，决定文件名
        yesterday = datetime.now() - timedelta(days=1)
        report_date_str = yesterday.strftime("%Y%m%d")

        for r in range(1, 4):
            for c in range(1, 5):
                val = str(raw_sheet.cell(r, c).value)
                if "生产运营情况" in val:
                    match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', val)
                    if match:
                        report_date_str = f"{match.group(1)}{int(match.group(2)):02d}{int(match.group(3)):02d}"
                    break

        output_filename = f"安全生产每日简报 {report_date_str}.xlsx"

        # 写入系统数据
        copy_system_data_sequential(raw_sheet, tmpl_sheet)

        # 定位中部网格
        r_zc = get_row_by_keywords(tmpl_sheet, ["增城建废", "增城"], 50)
        r_by = get_row_by_keywords(tmpl_sheet, ["白云建废", "白云"], r_zc + 10)
        r_sewage = get_row_by_keywords(tmpl_sheet, ["检查专项", "垃圾转运站污水", "污水转运"], r_by + 10)

        # 写入增城建废数据 (增加“再生石粉”)
        inject_manual_data(tmpl_sheet, ["装修垃圾"], z_zx, r_zc, r_by)
        inject_manual_data(tmpl_sheet, ["砖渣"], z_zz, r_zc, r_by)
        inject_manual_data(tmpl_sheet, ["混凝土"], z_hnt, r_zc, r_by)
        inject_manual_data(tmpl_sheet, ["再生水稳"], z_sw, r_zc, r_by)
        inject_manual_data(tmpl_sheet, ["再生石粉"], z_sf, r_zc, r_by)

        # 写入白云建废数据
        inject_manual_data(tmpl_sheet, ["建筑废弃物进厂", "建废"], by_jianfei_total, r_by, r_sewage)
        inject_manual_data(tmpl_sheet, ["混凝土出厂量", "原生混凝土", "混凝土出厂"], by_hunningtu, r_by, r_sewage)
        inject_manual_data(tmpl_sheet, ["再生骨料1-2石", "再生1-2石", "再生骨料"], by_zaisheng12, r_by, r_sewage)
        inject_manual_data(tmpl_sheet, ["再生水稳"], by_zaishengshuiwen, r_by, r_sewage)

        # 写入表尾文字段落 (注入检查项)
        fix_bottom_paragraphs(raw_sheet, tmpl_sheet, tot_sewage, to_hf, to_pwr, hf_data, checks)

        # 9. 保存至内存
        out = io.BytesIO()
        tmpl_wb.save(out)
        out.seek(0)

        # 缓存生成的文件
        file_id = str(uuid.uuid4())
        
        # 内存溢出防御机制：当缓存大于15个时，清理最早的5个
        if len(GENERATED_FILES) > 15:
            sorted_keys = list(GENERATED_FILES.keys())
            for k in sorted_keys[:5]:
                del GENERATED_FILES[k]

        GENERATED_FILES[file_id] = {
            "filename": output_filename,
            "data": out
        }

        # 10. 提取的解析数据预览返回给前端
        parsed_preview = {
            "sewage": {
                "total": tot_sewage,
                "to_hf": to_hf,
                "to_power": to_pwr
            },
            "baiyun": {
                "shizha": by_shizha,
                "shuinitou": by_shuinitou,
                "yuanshengshi": by_yuanshengshi,
                "total": by_jianfei_total,
                "hunningtu": by_hunningtu,
                "zaisheng12": by_zaisheng12,
                "zaishengshuiwen": by_zaishengshuiwen
            },
            "zengcheng": {
                "装修垃圾": z_zx,
                "再生水稳": z_sw,
                "再生石粉": z_sf,
                "砖渣": z_zz,
                "混凝土块": z_hnt
            },
            "huanfu": hf_data
        }

        return jsonify({
            'success': True,
            'file_id': file_id,
            'filename': output_filename,
            'check_result': {
                'status': check_status,
                'message': check_message
            },
            'parsed_data': parsed_preview
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f"生成日报失败: {str(e)}"}), 500


@app.route('/api/download/<file_id>', methods=['GET'])
def download_file(file_id):
    file_info = GENERATED_FILES.get(file_id)
    if not file_info:
        return "<h3>文件已过期或不存在，请重新生成！</h3>", 404

    file_info["data"].seek(0)
    
    # 兼容 Flask 1.x 与 Flask 2.x
    import flask
    flask_version = flask.__version__
    kwargs = {
        "as_attachment": True,
        "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    
    if flask_version.startswith('1.'):
        kwargs["attachment_filename"] = file_info["filename"]
        kwargs["cache_timeout"] = 0
    else:
        kwargs["download_name"] = file_info["filename"]
        
    return send_file(file_info["data"], **kwargs)


if __name__ == '__main__':
    # 默认 5000 端口
    app.run(host='0.0.0.0', port=5000, debug=True)
